"""
guardar_resultados.py
Persistencia de resultados en un archivo Excel (`resultados.xlsx`).

Se crean (si no existen) dos hojas: 'resumen' (una fila por ejecución)
y 'detalles' (una fila por pregunta vinculada a la ejecución mediante `run_id`).
"""

import os
from datetime import datetime

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None

RESULTADOS_FILE_XLSX = "resultados.xlsx"


def _iso_ts():
    return datetime.now().isoformat(timespec="seconds")


def guardar_resultados(curp, respuestas, omitidos, resultados_globales, total_time_seconds=None, ubicacion=""):
    """Guarda los resultados en `resultados.xlsx`.

    Args:
        curp (str): identificador del caso
        respuestas (dict): diccionario con las respuestas por pregunta
        omitidos (int): número de preguntas omitidas
        resultados_globales (dict): porcentajes por categoría
        total_time_seconds (float): duración total del formulario en segundos
        ubicacion (str): texto libre indicando dónde se realizó el formulario
    """
    if Workbook is None:
        # Si no está openpyxl, caemos a un CSV simple para evitar excepción inesperada.
        import csv
        csv_file = "resultados_fallback.csv"
        existe = os.path.exists(csv_file)
        with open(csv_file, mode="a", newline="", encoding="utf-8") as archivo:
            escritor = csv.writer(archivo)
            if not existe:
                escritor.writerow(["timestamp", "curp", "pregunta_id", "pregunta", "respuesta", "valor", "omitida", "omitidos", "ubicacion", "total_time_seconds"])
            timestamp = _iso_ts()
            for pregunta_id, datos in respuestas.items():
                escritor.writerow([
                    timestamp,
                    curp,
                    pregunta_id,
                    datos.get("pregunta", ""),
                    datos.get("respuesta", ""),
                    datos.get("valor", 0.0),
                    datos.get("omitida", False),
                    omitidos,
                    ubicacion,
                    f"{(total_time_seconds or 0):.1f}"
                ])
        return

    file_exists = os.path.exists(RESULTADOS_FILE_XLSX)
    if file_exists:
        wb = load_workbook(RESULTADOS_FILE_XLSX)
    else:
        wb = Workbook()

    # Hoja resumen
    if "resumen" not in wb.sheetnames:
        resumen = wb.create_sheet("resumen")
        resumen.append([
            "run_id",
            "timestamp",
            "curp",
            "ubicacion",
            "total_time_seconds",
            "omitidos",
            "violencia_fisica",
            "violencia_psicologica",
            "violencia_economica",
            "violencia_sexual",
            "violencia_verbal",
            "violencia_digital",
            "negligencia",
        ])
    else:
        resumen = wb["resumen"]

    # Hoja detalles
    if "detalles" not in wb.sheetnames:
        detalles = wb.create_sheet("detalles")
        detalles.append([
            "run_id",
            "timestamp",
            "curp",
            "pregunta_id",
            "pregunta",
            "respuesta",
            "valor",
            "omitida",
        ])
    else:
        detalles = wb["detalles"]

    # Generar run_id y escribir resumen
    timestamp = _iso_ts()
    run_id = timestamp.replace(":", "-")

    viol_fisica = resultados_globales.get("violencia_fisica", 0.0)
    viol_psico = resultados_globales.get("violencia_psicologica", 0.0)
    viol_econ = resultados_globales.get("violencia_economica", 0.0)
    viol_sex = resultados_globales.get("violencia_sexual", 0.0)
    viol_ver = resultados_globales.get("violencia_verbal", 0.0)
    viol_dig = resultados_globales.get("violencia_digital", 0.0)
    negligencia = resultados_globales.get("negligencia", 0.0)

    resumen.append([
        run_id,
        timestamp,
        curp,
        ubicacion,
        float(total_time_seconds or 0.0),
        int(omitidos),
        float(viol_fisica),
        float(viol_psico),
        float(viol_econ),
        float(viol_sex),
        float(viol_ver),
        float(viol_dig),
        float(negligencia),
    ])

    # Escribir detalles por pregunta
    for pregunta_id, datos in respuestas.items():
        detalles.append([
            run_id,
            timestamp,
            curp,
            pregunta_id,
            datos.get("pregunta", ""),
            datos.get("respuesta", ""),
            float(datos.get("valor", 0.0)),
            bool(datos.get("omitida", False)),
        ])

    # Guardar workbook
    # Si workbook fue creado con hoja por defecto vacía, eliminarla
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 2:
        try:
            std = wb["Sheet"]
            wb.remove(std)
        except Exception:
            pass

    wb.save(RESULTADOS_FILE_XLSX)

